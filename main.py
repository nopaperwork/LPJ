# Impor the dependencies
# from hashlib import new
# from unittest.main import MAIN_EXAMPLES

from ast import main
from pickle import FALSE
import requests
import bs4
import sys
from xlsxwriter import Workbook
from selenium import webdriver
from datetime import datetime
import sqlite3 as sl
import time
import os
from selenium.common.exceptions import WebDriverException
import json
import re
import openpyxl
from pathlib import Path
from config import DIFF_DAYS
from config import INPUT_FOLDER
from config import PAYLOAD_FILE
from config import HELP_FILE
from config import OUTPUT_HTML_FOLDER
from config import OUTPUT_EXCEL
from config import OUTPUT_PDF_FOLDER
from config import CHROME_PATH
from config import DIFF_DAYS

# Start the scraping
url = (
    "https://www.indiapost.gov.in/_layouts/15/DOP.Portal.Tracking/TrackConsignment.aspx"
)

# Get the payload
final_value_list = []
with open(PAYLOAD_FILE, "r") as f:
    payload = f.readline()
# Read the excel file for input from the INPUT FOLDER
file_list = os.listdir(INPUT_FOLDER)
tracker_dict = {}
tracker_keys = ()
for file in file_list:
    # Loop through each file
    file = INPUT_FOLDER + file
    if file.endswith(".xlsx"):
        book = openpyxl.load_workbook(file)
        print(file)
        try:
            SHEET_NAME = file.split('--')[-1].strip('.xlsx')
            sheet = book[SHEET_NAME]
        except:
            print('Issue with File naming convention for file' , str(file))
            sys.exit()
        # here you iterate over the rows in the specific column
        for row in range(5, sheet.max_row + 1):
            # for row in range(2,6):
            # SL/ARTICLE NO/REF NO/CITY/PIN CODE/District/NAME/ADD 1/ADD 2/ADD 3/WEIGHT
            main_key = red_id  = destination_pin_code = user_note = addressee = addressee_address3 = addressee_address1 = addressee_address2 = addressee_address= None
            for column in "CDEHIJKM":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                if sheet[cell_name].value:
                    # Read the Main key in Column A
                    if column == "C":
                        main_key = sheet[cell_name].value.strip()
                    # Read the Refid from column B
                    if column == "D":
                        red_id = sheet[cell_name].value.strip()
                    # Read the source PIN in column C
                    # if column == "D":
                    #     booked_city = str(sheet[cell_name].value).strip()
                    # Read the target PIN in column D
                    if column == "E":
                        destination_pin_code = str(sheet[cell_name].value).strip()
                    # Read the target PIN in column D
                    # if column == "F":
                    #     booked_pin_code = str(sheet[cell_name].value).strip()
                    # Read the target PIN in column D
                    if column == "H":
                        addressee = str(sheet[cell_name].value).strip()
                    # Read the target PIN in column D
                    if column == "I":
                        addressee_address1 = str(sheet[cell_name].value).strip()
                    if column == "J":
                        addressee_address2 = str(sheet[cell_name].value).strip()
                    if column == "K":
                        addressee_address3 = str(sheet[cell_name].value).strip()
                    if column == "M":
                        user_note = str(sheet[cell_name].value).strip()
            booked_date = sheet['G2'].value.strip()
            if main_key:
                if addressee_address1 == None :
                    addressee_address1 = ''
                if addressee_address2 == None :
                    addressee_address2 = ''
                if addressee_address3 == None :
                    addressee_address3 = ''
                addressee_address = addressee_address1 + '/' + addressee_address2 + '/' + addressee_address3   
                # If main key is not None , appned it in the dictionary
                tracker_dict[main_key] = [red_id, booked_date ,destination_pin_code,addressee ,addressee_address,user_note]                
                
# print(tracker_dict)

for key in tracker_dict.keys():
    tracker_keys += (key,)
# print(tracker_keys)

# Remove the files from the HTML folder
for entry in Path(OUTPUT_HTML_FOLDER).iterdir():
    if entry.is_file():
        print("Removed:", entry)
        entry.unlink()

# Check the DB for the keys 
# Find out the entries where status is not set
con = sl.connect('dont_touch.db')
with con:
    # data = con.execute("SELECT * FROM central_tracker where delivery_status != '' and article_number IN " + str(tracker_keys))
    data = con.execute("SELECT * FROM central_tracker ")
    for i ,row in enumerate(data):
        print(row)
        if row[10].strip() != '' : 
            tracker_dict.pop(row[0], None)
# con.close()
# # Now we have the entries which is not updated in central tracker
# input()
# sys.exit()
# Loop over the dictionary

# Plan is to save it as the HTML files by changing the model file
# Get the model file
with open(HELP_FILE, "r", encoding="utf-8") as f:
    html = f.read()

print(tracker_dict)
for main_key, value_list in tracker_dict.items():
    print(main_key)
    # Replace the ID in the string
    to_be_string = (
        "&ctl00%24PlaceHolderMain%24ucNewLegacyControl%24txtOrignlPgTranNo=" + main_key
    )
    replaced = re.sub(
        "&ctl00%24PlaceHolderMain%24ucNewLegacyControl%24txtOrignlPgTranNo=[^&]*",
        to_be_string,
        payload,
    )

    # Headers
    headers = {
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Cookie": "WSS_FullScreenMode=false",
        "Origin": "https://www.indiapost.gov.in",
        "Referer": "https://www.indiapost.gov.in/_layouts/15/DOP.Portal.Tracking/TrackConsignment.aspx",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "Sec-GPC": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.114 Safari/537.36",
        "X-MicrosoftAjax": "Delta=true",
        "X-Requested-With": "XMLHttpRequest",
    }
    not_continue = False
    counter = 0
    while True:
        try:
            # Try to ge the response
            response = requests.request("POST", url, headers=headers, data=replaced)
            break
        except:
            print('Trying Again')
            counter += 1
            if counter >50 :
                print('No response for after 50 attempts:' ,main_key)
                not_continue = True
                break
            continue
    
    # Break the loop
    if not_continue:
        continue
        
    # Soup the response
    soup = bs4.BeautifulSoup(response.text, "html.parser")
    # Select the new content from the response (main table with data)
    row_div = soup.select(
        'div[id="ctl00_PlaceHolderMain_ucNewLegacyControl_divTrckMailArticleOER"]'
    )
    if row_div :
        new_content = str(row_div[0])
    else:
        print('No data found for :', main_key)
        value_list.append('NA')
        value_list.insert(0, main_key)
        final_value_list.append(value_list)
        continue
    
    # Soup the help file and replace it with new content
    soup = bs4.BeautifulSoup(html, "html.parser")
    outer_div = soup.find(
        "div",
        attrs={"id": "ctl00_PlaceHolderMain_ucNewLegacyControl_divTrckMailArticleOER"},
    )
    outer_div.clear()
    outer_div.append(bs4.BeautifulSoup(new_content, "html.parser"))

    # Do the same for Second Content (Header inbox)
    soup_second = bs4.BeautifulSoup(response.text, "html.parser")
    row_div_second = soup_second.select(
        'div[id="ctl00_PlaceHolderMain_ucNewLegacyControl_divTrckConsgHomePg"]'
    )
    new_content_second = str(row_div_second[0])
    outer_div = soup.find(
        "div",
        attrs={"id": "ctl00_PlaceHolderMain_ucNewLegacyControl_divTrckConsgHomePg"},
    )
    outer_div.clear()
    outer_div.append(bs4.BeautifulSoup(new_content_second, "html.parser"))

    # Form the File name
    # file_name = str(value_list[0]).replace("/", "_") + "_" + str(main_key) + ".html"
    file_name =   str(value_list[0]).replace("/", "_") + "_" + str(main_key) +  "_" + str(value_list[-1]) + ".html"
    file_name_list = file_name.split("_")
    file_name = OUTPUT_HTML_FOLDER + "_".join(file_name_list[2:])

    print(file_name)
    with open(file_name, "w", encoding="utf-8") as file:
        file.write(str(soup))
    # Now we have the html file saved , get the data to update in excel
    # data will be in row_div
    data = []
    table = row_div[0].find(
        "table", attrs={"class": "responsivetable MailArticleEvntOER"}
    )
    rows = table.find_all("tr")
    for row in rows:
        cols = row.find_all("td")
        cols = [ele.text.strip() for ele in cols]
        if len(cols) >= 4:
            data.append(cols)
    # Now determine the date and check the status only if the date is older than 15 days
    data_date = []
    table_date = row_div[0].find(
        "table", attrs={"class": "responsivetable MailArticleOER"}
    )
    rows_date = table_date.find_all("tr")
    for row_date in rows_date:
        cols_date = row_date.find_all("td")
        cols_date = [ele.text.strip() for ele in cols_date]
        if len(cols_date) > 4:
            # print(cols_date)
            date_to_compare = cols_date[1]
            try :
                datetime_object = datetime.strptime(date_to_compare, "%d/%m/%Y %H:%M:%S")
                diff = datetime.now() - datetime_object
                print("Days past post booking :", diff.days)
                delivery_location = cols_date[-2]
                print("Delivery location", delivery_location)
                booked_at_location = cols_date[0]
                print("Booked at", booked_at_location)
            except:
                print("Different Format encountered")

    try: 
        print(diff)
    except :
        print('No data found for :', main_key)
        value_list.append('NA')
        value_list.insert(0, main_key)
        final_value_list.append(value_list)
        continue
        
    if diff.days >= DIFF_DAYS:
        # Determine status based on the priority
        status_found = False
        for event_entry in data:
            # print(event_entry[-1].upper())
            if "REFUSED" in event_entry[-1].upper():
                # print('REFUSED')
                value_list.append("Item Refused")
                status_found = True
                break
        if not status_found:
            for event_entry in data:
                if (
                    "UNCLAIMED" in event_entry[-1].upper()
                    or "RETURN" in event_entry[-1].upper()
                    or "RETURNED" in event_entry[-1].upper()
                    or "INSUFFICIENT" in event_entry[-1].upper()
                    or "NO SUCH PERSON" in event_entry[-1].upper()
                ):
                    value_list.append("Item Not Delivered")
                    status_found = True
                    break
        if not status_found:
            new_data = [data[0], data[-1]]
            # Check the first entry and last entry SO to determine status
            # Updated logic
            if new_data[0][-1].upper() == "ITEM DELIVERY CONFIRMED" or (
                "ITEM DELIVERED" in new_data[0][-1].upper()
            ):
                if booked_at_location != delivery_location:
                    value_list.append("Delivered")
                    status_found = True
                else:
                    value_list.append("Manually Check")
                    status_found = True
        if not status_found:
            value_list.append("Item Delivered")
    else:
        # Append Space value
        value_list.append("") 
    value_list.insert(0, main_key)
    final_value_list.append(value_list)

print(final_value_list)

# Here we have the data available now
# Push it to the DB
sql = """INSERT INTO central_tracker (article_number, ref_id , booking_date, destination_pincode,addressee,addressee_address 
                        , user_note , delivery_status) values(?, ?,?, ? ,? , ?, ?, ?)"""
data = final_value_list
with con:
    con.executemany(sql, data)
con.close()

# update the excel tracker -temporary one
# Header of the Excel
ordered_list = [
    "ARTICLE NO",
    "REF NO",
    "Booking date",
    "Delivery PIN CODE",
    "Addressee",
    "Addressee Address",
    "User Note",
    "Delivery Status",
]  # List object calls by index, but the dict object calls items randomly
# Preapre the file name
filename = OUTPUT_EXCEL
filename = OUTPUT_EXCEL + 'excel_output' + str(time.strftime("%Y_%m_%d %H-%M-%S", time.gmtime())) +'.xlsx'
wb = Workbook(filename)
ws = wb.add_worksheet(
    time.strftime("%Y_%m_%d %H-%M-%S", time.gmtime())
)  # Or leave it blank. The default name is "Sheet 1"

first_row = 0
# push the header information
for header in ordered_list:
    col = ordered_list.index(header)  # We are keeping order.
    ws.write(
        first_row, col, header
    )  # We have written first row which is the header of worksheet also.
row = 1
# Push the data to excel file
for player in final_value_list:
    col = 0
    ws.write_row(row, col, player)
    row += 1  # enter the next row
# Close the work book
wb.close()
print("Finished scraping for  Post information !")

# Create Output folder if does not exist
OUTPUT_PDF_FOLDER = OUTPUT_PDF_FOLDER + 'pdf_output' + str(time.strftime("%Y_%m_%d", time.gmtime())) + '/'
path = OUTPUT_PDF_FOLDER
# Check whether the specified path exists or not
isExist = os.path.exists(path)
if not isExist:
    # Create a new directory because it does not exist 
    os.makedirs(path)
    print("The new directory is created!")

# Now convert the html files to PDF
chrome_options = webdriver.ChromeOptions()
settings = {
    "recentDestinations": [
        {
            "id": "Save as PDF",
            "origin": "local",
            "account": "",
        }
    ],
    "selectedDestinationId": "Save as PDF",
    "version": 2,
}
output_pdf_path = os.path.abspath(OUTPUT_PDF_FOLDER)
prefs = {
    "printing.print_preview_sticky_settings.appState": json.dumps(settings),
    "savefile.default_directory": output_pdf_path,
}
# 'savefile.default_directory':  OUTPUT_PDF_FOLDER}


file_list = os.listdir(OUTPUT_HTML_FOLDER)
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument("--kiosk-printing")
driver = webdriver.Chrome(executable_path=CHROME_PATH, chrome_options=chrome_options)

for file in file_list:
    if file.endswith(".html"):
        output_file = os.path.abspath(OUTPUT_HTML_FOLDER + file)
        filename = "file://" + output_file
        driver.get(filename)
        try:
            os.remove(OUTPUT_PDF_FOLDER + "Track Consignment.pdf")
        except:
            pass
        # Print the file
        driver.execute_script("window.print();")

        # Print the new file
        new_filename = OUTPUT_PDF_FOLDER + file.replace(".html", ".pdf")
        time.sleep(0.5)
        try:
            os.rename(OUTPUT_PDF_FOLDER + "Track Consignment.pdf", new_filename)
        except FileExistsError:
            print("File already exist for :", new_filename)
            print("Updating with latest info")
            os.remove(new_filename)
            os.rename(OUTPUT_PDF_FOLDER + "Track Consignment.pdf", new_filename)
        except FileNotFoundError:
            print("Folder not found for" , new_filename) 
        except :
            print('cannot save the file for :',OUTPUT_PDF_FOLDER + "Track Consignment.pdf")
driver.close()
print("Completed")
input()
